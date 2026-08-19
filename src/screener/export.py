from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.screener.engine import load_financial_data, run_screener

OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "screener_output.xlsx"


PRESETS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch",
]


# ---------------------------------------------------------------------
# SCORE COMPONENTS
# ---------------------------------------------------------------------

SCORE_METRICS = {
    "Profitability": {
        "return_on_equity_pct": 15,
        "return_on_capital_employed_pct": 10,
        "net_profit_margin_pct": 10,
    },
    "Cash Quality": {
        "fcf_cagr": 15,
        "cfo_pat_ratio": 10,
        "fcf_positive_flag": 5,
    },
    "Growth": {
        "revenue_cagr_5yr": 10,
        "pat_cagr_5yr": 10,
    },
    "Leverage": {
        "debt_to_equity": 10,
        "interest_coverage": 5,
    },
}


# ---------------------------------------------------------------------
# DATA PREPARATION
# ---------------------------------------------------------------------


def add_missing_score_columns(df):
    """
    Add score inputs that may not yet exist in the current
    financial_ratios table.

    Missing values are represented as NaN rather than causing
    the export to fail.
    """

    result = df.copy()

    defaults = {
        "return_on_capital_employed_pct": np.nan,
        "fcf_cagr": np.nan,
        "cfo_pat_ratio": np.nan,
    }

    for column, value in defaults.items():
        if column not in result.columns:
            result[column] = value

    result["fcf_positive_flag"] = (
        pd.to_numeric(
            result["free_cash_flow_cr"],
            errors="coerce",
        )
        > 0
    ).astype(float)

    return result


# ---------------------------------------------------------------------
# P10 / P90 WINSORIZATION
# ---------------------------------------------------------------------


def winsorize_series(series):
    """
    Cap values at the 10th and 90th percentile.
    """

    values = pd.to_numeric(
        series,
        errors="coerce",
    )

    valid = values.dropna()

    if valid.empty:
        return values

    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)

    return values.clip(
        lower=p10,
        upper=p90,
    )


def normalize_0_100(series, inverse=False):
    """
    Normalize a metric to 0-100 after P10/P90 winsorization.

    inverse=True means lower values are better.
    """

    capped = winsorize_series(series)

    valid = capped.dropna()

    if valid.empty:
        return pd.Series(
            50.0,
            index=series.index,
        )

    minimum = valid.min()
    maximum = valid.max()

    if maximum == minimum:
        result = pd.Series(
            50.0,
            index=series.index,
        )

    else:
        result = ((capped - minimum) / (maximum - minimum)) * 100

    if inverse:
        result = 100 - result

    return result.fillna(50.0)


# ---------------------------------------------------------------------
# SECTOR-RELATIVE NORMALIZATION
# ---------------------------------------------------------------------


def sector_normalize(
    df,
    column,
    inverse=False,
):
    """
    Normalize a metric within broad_sector.

    This ensures the composite score reflects performance
    relative to sector peers.
    """

    result = pd.Series(
        np.nan,
        index=df.index,
        dtype=float,
    )

    if "broad_sector" not in df.columns:
        return normalize_0_100(
            df[column],
            inverse=inverse,
        )

    for sector, index in df.groupby(
        "broad_sector",
        dropna=False,
    ).groups.items():

        subset = df.loc[index, column]

        result.loc[index] = normalize_0_100(
            subset,
            inverse=inverse,
        )

    return result.fillna(50.0)


# ---------------------------------------------------------------------
# COMPOSITE QUALITY SCORE
# ---------------------------------------------------------------------


def calculate_composite_quality_score(df):
    """
    Calculate the 0-100 composite quality score.

    Weighting:

    Profitability = 35%
        ROE 15%
        ROCE 10%
        NPM 10%

    Cash Quality = 30%
        FCF CAGR 15%
        CFO/PAT 10%
        FCF positive 5%

    Growth = 20%
        Revenue CAGR 10%
        PAT CAGR 10%

    Leverage = 15%
        D/E 10%
        ICR 5%
    """

    result = add_missing_score_columns(df)

    # Sector-relative normalized metrics.

    roe_score = sector_normalize(
        result,
        "return_on_equity_pct",
    )

    roce_score = sector_normalize(
        result,
        "return_on_capital_employed_pct",
    )

    npm_score = sector_normalize(
        result,
        "net_profit_margin_pct",
    )

    fcf_cagr_score = sector_normalize(
        result,
        "fcf_cagr",
    )

    cfo_pat_score = sector_normalize(
        result,
        "cfo_pat_ratio",
    )

    fcf_positive_score = result["fcf_positive_flag"] * 100

    revenue_score = sector_normalize(
        result,
        "revenue_cagr_5yr",
    )

    pat_score = sector_normalize(
        result,
        "pat_cagr_5yr",
    )

    debt_score = sector_normalize(
        result,
        "debt_to_equity",
        inverse=True,
    )

    icr_score = sector_normalize(
        result,
        "interest_coverage",
    )

    # Weighted components.

    profitability = roe_score * 0.15 + roce_score * 0.10 + npm_score * 0.10

    cash_quality = (
        fcf_cagr_score * 0.15 + cfo_pat_score * 0.10 + fcf_positive_score * 0.05
    )

    growth = revenue_score * 0.10 + pat_score * 0.10

    leverage = debt_score * 0.10 + icr_score * 0.05

    result["composite_quality_score"] = (
        profitability + cash_quality + growth + leverage
    ).clip(
        lower=0,
        upper=100,
    )

    return result


# ---------------------------------------------------------------------
# PRESET THRESHOLD CHECK
# ---------------------------------------------------------------------


def threshold_columns_for_preset(
    preset,
):
    """
    Return the threshold rules used for Excel highlighting.
    """

    rules = {
        "quality_compounder": {
            "return_on_equity_pct": ("min", 15),
            "debt_to_equity": ("max", 1),
            "free_cash_flow_cr": ("min", 0),
            "revenue_cagr_5yr": ("min", 10),
        },
        "value_pick": {
            "pe": ("max", 20),
            "pb": ("max", 3),
            "debt_to_equity": ("max", 2),
            "dividend_yield": ("min", 1),
        },
        "growth_accelerator": {
            "pat_cagr_5yr": ("min", 20),
            "revenue_cagr_5yr": ("min", 15),
            "debt_to_equity": ("max", 2),
        },
        "dividend_champion": {
            "dividend_yield": ("min", 2),
            "dividend_payout_ratio_pct": ("max", 80),
            "free_cash_flow_cr": ("min", 0),
        },
        "debt_free_blue_chip": {
            "debt_to_equity": ("max", 0),
            "return_on_equity_pct": ("min", 12),
            "sales": ("min", 5000),
        },
        "turnaround_watch": {
            "revenue_cagr_3yr": ("min", 10),
            "free_cash_flow_cr": ("min", 0),
        },
    }

    return rules.get(
        preset,
        {},
    )


def threshold_passes(
    value,
    rule,
):
    if pd.isna(value):
        return False

    operator, threshold = rule

    if operator == "min":
        return value >= threshold

    if operator == "max":
        return value <= threshold

    return False


# ---------------------------------------------------------------------
# EXCEL FORMATTING
# ---------------------------------------------------------------------

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)


def format_sheet(
    workbook,
    sheet_name,
    preset,
):
    ws = workbook[sheet_name]

    headers = [cell.value for cell in ws[1]]

    header_index = {value: index + 1 for index, value in enumerate(headers)}

    rules = threshold_columns_for_preset(preset)

    for metric, rule in rules.items():

        if metric not in header_index:
            continue

        column_number = header_index[metric]

        for row in range(
            2,
            ws.max_row + 1,
        ):

            cell = ws.cell(
                row=row,
                column=column_number,
            )

            try:
                value = float(cell.value)
            except (
                TypeError,
                ValueError,
            ):
                cell.fill = RED_FILL
                continue

            if threshold_passes(
                value,
                rule,
            ):
                cell.fill = GREEN_FILL
            else:
                cell.fill = RED_FILL

    # Header formatting.
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    # Auto width.
    for column_cells in ws.columns:

        maximum = 0

        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            if cell.value is not None:
                maximum = max(
                    maximum,
                    len(str(cell.value)),
                )

        ws.column_dimensions[column_letter].width = min(
            maximum + 2,
            30,
        )


# ---------------------------------------------------------------------
# EXPORT
# ---------------------------------------------------------------------


def export_screener_output():
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    base_data = load_financial_data()

    scored_data = calculate_composite_quality_score(base_data)

    # Keep the requested KPI-oriented columns.
    export_columns = [
        "company_id",
        "year",
        "broad_sector",
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "debt_to_equity",
        "interest_coverage",
        "free_cash_flow_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "asset_turnover",
        "earnings_per_share",
        "book_value_per_share",
        "dividend_payout_ratio_pct",
        "sales",
        "net_profit",
        "composite_quality_score",
    ]

    # Only export columns actually available.
    export_columns = [
        column for column in export_columns if column in scored_data.columns
    ]

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        for preset in PRESETS:

            result = run_screener(preset=preset)

            # Merge composite score.
            result = result.drop(
                columns=["composite_quality_score"],
                errors="ignore",
            )

            result = result.merge(
                scored_data[
                    [
                        "company_id",
                        "composite_quality_score",
                    ]
                ],
                on="company_id",
                how="left",
            )

            result = result.sort_values(
                "composite_quality_score",
                ascending=False,
                na_position="last",
            )

            columns = [column for column in export_columns if column in result.columns]

            result[columns].to_excel(
                writer,
                sheet_name=preset[:31],
                index=False,
            )

    # Apply cell-level formatting.
    workbook = load_workbook(OUTPUT_FILE)

    for preset in PRESETS:

        sheet_name = preset[:31]

        if sheet_name not in workbook.sheetnames:
            continue

        format_sheet(
            workbook,
            sheet_name,
            preset,
        )

    workbook.save(OUTPUT_FILE)

    return OUTPUT_FILE


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

if __name__ == "__main__":

    print("=" * 70)
    print("DAY 17 — COMPOSITE SCORE + EXCEL EXPORT")
    print("=" * 70)

    output = export_screener_output()

    print()
    print(
        "Output:",
        output.resolve(),
    )

    workbook = load_workbook(
        output,
        read_only=True,
    )

    print(
        "Sheets:",
        workbook.sheetnames,
    )

    print()
    print("Sheet counts:")

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        print(f"{sheet:25s}: " f"{max(ws.max_row - 1, 0):3d} rows")

    print()
    print("✅ DAY 17 COMPLETE")
