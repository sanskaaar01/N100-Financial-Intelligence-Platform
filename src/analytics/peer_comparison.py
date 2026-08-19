import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DB_PATH = Path("db/nifty100.db")
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "peer_comparison.xlsx"


METRICS = [
    ("ROE", "ROE"),
    ("ROCE", "ROCE"),
    ("Net Profit Margin", "Net Profit Margin"),
    ("D/E", "D/E"),
    ("FCF", "FCF"),
    ("PAT CAGR 5yr", "PAT CAGR 5yr"),
    ("Revenue CAGR 5yr", "Revenue CAGR 5yr"),
    ("EPS CAGR 5yr", "EPS CAGR 5yr"),
    ("Interest Coverage", "Interest Coverage"),
    ("Asset Turnover", "Asset Turnover"),
]


GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

BENCHMARK_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

MEDIAN_FILL = PatternFill(
    fill_type="solid",
    fgColor="E7E6E6",
)


def get_connection():
    return sqlite3.connect(DB_PATH)


def load_peer_groups():
    conn = get_connection()

    try:
        return pd.read_sql_query(
            """
            SELECT
                peer_group_name,
                company_id,
                is_benchmark
            FROM peer_groups
            ORDER BY peer_group_name, company_id
            """,
            conn,
        )
    finally:
        conn.close()


def load_peer_percentiles():
    conn = get_connection()

    try:
        return pd.read_sql_query(
            """
            SELECT
                company_id,
                peer_group_name,
                metric,
                value,
                percentile_rank,
                year
            FROM peer_percentiles
            """,
            conn,
        )
    finally:
        conn.close()


def load_company_names():
    conn = get_connection()

    try:
        schema = pd.read_sql_query(
            "PRAGMA table_info(companies)",
            conn,
        )

        columns = schema["name"].tolist()

        id_candidates = [
            "company_id",
            "id",
            "symbol",
            "company",
            "name",
        ]

        name_candidates = [
            "company_name",
            "name",
            "company",
            "stock_name",
            "title",
        ]

        id_column = None

        for candidate in id_candidates:
            if candidate in columns:
                id_column = candidate
                break

        name_column = None

        for candidate in name_candidates:
            if candidate in columns:
                name_column = candidate
                break

        if id_column is None:
            return pd.DataFrame(
                columns=[
                    "company_id",
                    "company_name",
                ]
            )

        if name_column is None:
            query = f"""
                SELECT
                    "{id_column}" AS company_id
                FROM companies
            """

            result = pd.read_sql_query(
                query,
                conn,
            )

            result["company_name"] = result["company_id"]

            return result[
                [
                    "company_id",
                    "company_name",
                ]
            ]

        query = f"""
            SELECT
                "{id_column}" AS company_id,
                "{name_column}" AS company_name
            FROM companies
        """

        result = pd.read_sql_query(
            query,
            conn,
        )

        return result[
            [
                "company_id",
                "company_name",
            ]
        ]

    finally:
        conn.close()


def build_comparison_table(
    peer_percentiles,
    peer_groups,
    company_names,
):
    if peer_percentiles.empty:
        return pd.DataFrame()

    rows = []

    for (
        peer_group_name,
        company_id,
    ), group in peer_percentiles.groupby(
        [
            "peer_group_name",
            "company_id",
        ],
        dropna=False,
    ):

        row = {
            "peer_group_name": peer_group_name,
            "company_id": company_id,
        }

        years = group["year"].dropna()

        if not years.empty:
            row["year"] = years.iloc[0]
        else:
            row["year"] = None

        for metric_name, _ in METRICS:

            metric_rows = group[group["metric"] == metric_name]

            if metric_rows.empty:

                row[metric_name] = np.nan

                row[f"{metric_name}_percentile"] = np.nan

            else:

                metric_row = metric_rows.iloc[0]

                row[metric_name] = pd.to_numeric(
                    metric_row["value"],
                    errors="coerce",
                )

                percentile = pd.to_numeric(
                    metric_row["percentile_rank"],
                    errors="coerce",
                )

                if pd.isna(percentile):
                    row[f"{metric_name}_percentile"] = np.nan
                else:
                    row[f"{metric_name}_percentile"] = float(percentile) * 100

        rows.append(row)

    result = pd.DataFrame(rows)

    if result.empty:
        return result

    if not company_names.empty:

        names = company_names.drop_duplicates("company_id")

        result = result.merge(
            names,
            on="company_id",
            how="left",
        )

        result["company_name"] = result["company_name"].fillna(result["company_id"])

    else:

        result["company_name"] = result["company_id"]

    benchmark_lookup = peer_groups[peer_groups["is_benchmark"] == 1][
        [
            "peer_group_name",
            "company_id",
        ]
    ].drop_duplicates()

    benchmark_lookup["is_benchmark"] = True

    result = result.merge(
        benchmark_lookup,
        on=[
            "peer_group_name",
            "company_id",
        ],
        how="left",
    )

    result["is_benchmark"] = result["is_benchmark"].fillna(False).astype(bool)

    return result


def get_visible_columns():

    columns = [
        "company_id",
        "company_name",
        "year",
    ]

    for metric_name, _ in METRICS:
        columns.append(metric_name)

    for metric_name, _ in METRICS:
        columns.append(f"{metric_name}_percentile")

    return columns


def safe_sheet_name(name):
    invalid = [
        "\\",
        "/",
        "*",
        "?",
        ":",
        "[",
        "]",
    ]

    result = str(name)

    for character in invalid:
        result = result.replace(
            character,
            "_",
        )

    return result[:31]


def add_peer_median(
    ws,
    dataframe,
):
    median_row_number = ws.max_row + 2

    ws.cell(
        row=median_row_number,
        column=1,
        value="PEER MEDIAN",
    )

    for column_number in range(
        1,
        ws.max_column + 1,
    ):

        header = ws.cell(
            row=1,
            column=column_number,
        ).value

        if header in [
            "company_id",
            "company_name",
            "year",
        ]:

            if column_number != 1:
                ws.cell(
                    row=median_row_number,
                    column=column_number,
                    value="",
                )

            continue

        values = []

        for row_number in range(
            2,
            ws.max_row + 1,
        ):

            value = ws.cell(
                row=row_number,
                column=column_number,
            ).value

            try:
                numeric = float(value)

                if not np.isnan(numeric):
                    values.append(numeric)

            except (
                TypeError,
                ValueError,
            ):
                continue

        if values:

            ws.cell(
                row=median_row_number,
                column=column_number,
                value=float(np.median(values)),
            )

    for column_number in range(
        1,
        ws.max_column + 1,
    ):

        cell = ws.cell(
            row=median_row_number,
            column=column_number,
        )

        cell.fill = MEDIAN_FILL

        cell.font = Font(bold=True)

    return median_row_number


def apply_excel_formatting(
    workbook,
    benchmark_map,
):
    for ws in workbook.worksheets:

        # -----------------------------
        # Header
        # -----------------------------

        for cell in ws[1]:

            cell.fill = HEADER_FILL

            cell.font = Font(bold=True)

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # -----------------------------
        # Find percentile columns
        # -----------------------------

        percentile_columns = []

        company_id_column = None

        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            header = ws.cell(
                row=1,
                column=column_number,
            ).value

            if header == "company_id":
                company_id_column = column_number

            if header is not None and str(header).endswith("_percentile"):
                percentile_columns.append(column_number)

        # -----------------------------
        # Percentile colour coding
        # -----------------------------

        for column_number in percentile_columns:

            for row_number in range(
                2,
                ws.max_row + 1,
            ):

                cell = ws.cell(
                    row=row_number,
                    column=column_number,
                )

                if cell.value is None:
                    continue

                try:
                    value = float(cell.value)
                except (
                    TypeError,
                    ValueError,
                ):
                    continue

                if value >= 75:

                    cell.fill = GREEN_FILL

                elif value >= 25:

                    cell.fill = YELLOW_FILL

                else:

                    cell.fill = RED_FILL

        # -----------------------------
        # Benchmark highlighting
        # -----------------------------

        if company_id_column is not None:

            benchmark_ids = benchmark_map.get(
                ws.title,
                set(),
            )

            for row_number in range(
                2,
                ws.max_row + 1,
            ):

                value = ws.cell(
                    row=row_number,
                    column=company_id_column,
                ).value

                if value is None:
                    continue

                if str(value) in benchmark_ids:

                    for column_number in range(
                        1,
                        ws.max_column + 1,
                    ):

                        ws.cell(
                            row=row_number,
                            column=column_number,
                        ).fill = BENCHMARK_FILL

        # -----------------------------
        # Median row
        # -----------------------------

        for row_number in range(
            2,
            ws.max_row + 1,
        ):

            value = ws.cell(
                row=row_number,
                column=1,
            ).value

            if value == "PEER MEDIAN":

                for column_number in range(
                    1,
                    ws.max_column + 1,
                ):

                    ws.cell(
                        row=row_number,
                        column=column_number,
                    ).fill = MEDIAN_FILL

                    ws.cell(
                        row=row_number,
                        column=column_number,
                    ).font = Font(bold=True)

        # -----------------------------
        # Freeze header
        # -----------------------------

        ws.freeze_panes = "A2"

        # -----------------------------
        # Auto width
        # -----------------------------

        for column_cells in ws.columns:

            maximum_length = 0

            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                if cell.value is None:
                    continue

                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)),
                )

            ws.column_dimensions[column_letter].width = min(
                maximum_length + 2,
                30,
            )


def write_excel(
    comparison,
    peer_groups,
):
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if OUTPUT_FILE.exists():
        OUTPUT_FILE.unlink()

    group_names = peer_groups["peer_group_name"].dropna().drop_duplicates().tolist()

    print()
    print(
        "Creating sheets:",
        len(group_names),
    )

    # Important:
    # Use openpyxl directly so that if one group has a
    # problem, we don't end up with a workbook with zero
    # visible sheets.

    from openpyxl import Workbook

    workbook = Workbook()

    # Remove default sheet.
    default_sheet = workbook.active

    workbook.remove(default_sheet)

    benchmark_map = {}

    for peer_group in group_names:

        sheet_name = safe_sheet_name(peer_group)

        print(f"Creating sheet: {peer_group}")

        group = comparison[comparison["peer_group_name"] == peer_group].copy()

        if group.empty:

            print(f"  WARNING: no rows for {peer_group}")

            # Still create the required sheet.
            ws = workbook.create_sheet(title=sheet_name)

            ws["A1"] = "No data available"

            continue

        # --------------------------------
        # Benchmark IDs
        # --------------------------------

        benchmark_rows = peer_groups[
            (peer_groups["peer_group_name"] == peer_group)
            & (peer_groups["is_benchmark"] == 1)
        ]

        benchmark_ids = set(benchmark_rows["company_id"].astype(str).tolist())

        benchmark_map[sheet_name] = benchmark_ids

        # --------------------------------
        # CRITICAL FIX:
        # Keep is_benchmark until AFTER sorting
        # --------------------------------

        group["is_benchmark"] = group["company_id"].astype(str).isin(benchmark_ids)

        group = group.sort_values(
            by=[
                "is_benchmark",
                "company_name",
                "company_id",
            ],
            ascending=[
                False,
                True,
                True,
            ],
        )

        visible_columns = get_visible_columns()

        visible_columns = [
            column for column in visible_columns if column in group.columns
        ]

        # --------------------------------
        # Create worksheet
        # --------------------------------

        ws = workbook.create_sheet(title=sheet_name)

        # Header
        for column_number, column_name in enumerate(
            visible_columns,
            start=1,
        ):

            cell = ws.cell(
                row=1,
                column=column_number,
                value=column_name,
            )

            cell.fill = HEADER_FILL

            cell.font = Font(bold=True)

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # --------------------------------
        # Company rows
        # --------------------------------

        for row_number, (_, row) in enumerate(
            group.iterrows(),
            start=2,
        ):

            for column_number, column_name in enumerate(
                visible_columns,
                start=1,
            ):

                value = row.get(
                    column_name,
                    np.nan,
                )

                if pd.isna(value):
                    value = None

                cell = ws.cell(
                    row=row_number,
                    column=column_number,
                    value=value,
                )

                # Percentile formatting
                if column_name.endswith("_percentile") and value is not None:

                    try:
                        percentile = float(value)

                        if percentile >= 75:
                            cell.fill = GREEN_FILL

                        elif percentile >= 25:
                            cell.fill = YELLOW_FILL

                        else:
                            cell.fill = RED_FILL

                    except (
                        TypeError,
                        ValueError,
                    ):
                        pass

        # --------------------------------
        # Benchmark row highlighting
        # --------------------------------

        company_id_index = None

        for index, column_name in enumerate(
            visible_columns,
            start=1,
        ):

            if column_name == "company_id":

                company_id_index = index

                break

        if company_id_index is not None:

            for row_number in range(
                2,
                ws.max_row + 1,
            ):

                company_id = ws.cell(
                    row=row_number,
                    column=company_id_index,
                ).value

                if company_id is not None and str(company_id) in benchmark_ids:

                    for column_number in range(
                        1,
                        ws.max_column + 1,
                    ):

                        ws.cell(
                            row=row_number,
                            column=column_number,
                        ).fill = BENCHMARK_FILL

        # --------------------------------
        # Peer median
        # --------------------------------

        median_row = ws.max_row + 2

        ws.cell(
            row=median_row,
            column=1,
            value="PEER MEDIAN",
        )

        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            column_name = ws.cell(
                row=1,
                column=column_number,
            ).value

            if column_name in [
                "company_id",
                "company_name",
                "year",
            ]:

                continue

            values = []

            for row_number in range(
                2,
                ws.max_row + 1,
            ):

                value = ws.cell(
                    row=row_number,
                    column=column_number,
                ).value

                try:

                    numeric = float(value)

                    if not np.isnan(numeric):
                        values.append(numeric)

                except (
                    TypeError,
                    ValueError,
                ):
                    pass

            if values:

                ws.cell(
                    row=median_row,
                    column=column_number,
                    value=float(np.median(values)),
                )

        # Median styling
        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            cell = ws.cell(
                row=median_row,
                column=column_number,
            )

            cell.fill = MEDIAN_FILL

            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"

        # --------------------------------
        # Width
        # --------------------------------

        for column_number in range(
            1,
            ws.max_column + 1,
        ):

            maximum = 0

            for row_number in range(
                1,
                ws.max_row + 1,
            ):

                value = ws.cell(
                    row=row_number,
                    column=column_number,
                ).value

                if value is not None:

                    maximum = max(
                        maximum,
                        len(str(value)),
                    )

            ws.column_dimensions[get_column_letter(column_number)].width = min(
                maximum + 2,
                30,
            )

    # --------------------------------
    # Safety check
    # --------------------------------

    if len(workbook.sheetnames) == 0:

        raise RuntimeError("No peer-group sheets were created.")

    workbook.save(OUTPUT_FILE)

    return workbook.sheetnames


def verify_output():
    if not OUTPUT_FILE.exists():

        raise FileNotFoundError(f"Output file was not created: {OUTPUT_FILE}")

    workbook = load_workbook(
        OUTPUT_FILE,
        read_only=True,
    )

    sheets = workbook.sheetnames

    print()
    print(
        "Output:",
        OUTPUT_FILE.resolve(),
    )

    print(
        "Sheets:",
        len(sheets),
    )

    print()

    for sheet_name in sheets:

        ws = workbook[sheet_name]

        company_rows = max(
            ws.max_row - 2,
            0,
        )

        print(f"{sheet_name:25s}: " f"{company_rows} company rows")

    if len(sheets) != 11:

        raise RuntimeError(f"Expected 11 sheets, got {len(sheets)}")

    print()
    print("✅ Exactly 11 peer-group sheets verified")

    return sheets


def run_peer_comparison():

    print("=" * 70)
    print("DAY 20 — PEER COMPARISON EXCEL REPORT")
    print("=" * 70)

    peer_groups = load_peer_groups()

    peer_percentiles = load_peer_percentiles()

    company_names = load_company_names()

    print()
    print(
        "Peer percentile rows:",
        len(peer_percentiles),
    )

    print(
        "Peer groups:",
        peer_groups["peer_group_name"].nunique(),
    )

    print(
        "Companies:",
        peer_groups["company_id"].nunique(),
    )

    print(
        "Company-name rows:",
        len(company_names),
    )

    comparison = build_comparison_table(
        peer_percentiles,
        peer_groups,
        company_names,
    )

    print()
    print(
        "Comparison rows:",
        len(comparison),
    )

    if comparison.empty:

        raise RuntimeError("Comparison table is empty.")

    sheets = write_excel(
        comparison,
        peer_groups,
    )

    verify_output()

    print()
    print("Sheet names:")

    for sheet in sheets:
        print(
            " ",
            sheet,
        )

    print()
    print("✅ DAY 20 PEER COMPARISON COMPLETE")


if __name__ == "__main__":
    run_peer_comparison()
