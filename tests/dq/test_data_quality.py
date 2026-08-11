import sqlite3
from pathlib import Path

from openpyxl import load_workbook


DB_PATH = Path("db/nifty100.db")
SCREENER_FILE = Path("output/screener_output.xlsx")
PEER_FILE = Path("output/peer_comparison.xlsx")
RADAR_DIR = Path("reports/radar_charts")


def get_connection():
    return sqlite3.connect(DB_PATH)


# ============================================================
# 1. DATABASE EXISTS
# ============================================================

def test_database_exists():
    assert DB_PATH.exists(), "SQLite database does not exist"


# ============================================================
# 2. FINANCIAL RATIOS TABLE EXISTS
# ============================================================

def test_financial_ratios_table_exists():
    conn = get_connection()

    try:
        result = conn.execute(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
              AND name = 'financial_ratios'
            """
        ).fetchone()
    finally:
        conn.close()

    assert result is not None


# ============================================================
# 3. FINANCIAL RATIOS ROW COUNT
# ============================================================

def test_financial_ratios_row_count():
    conn = get_connection()

    try:
        count = conn.execute(
            "SELECT COUNT(*) FROM financial_ratios"
        ).fetchone()[0]
    finally:
        conn.close()

    assert count >= 1100


# ============================================================
# 4. REQUIRED KPI COLUMNS EXIST
# ============================================================

def test_required_kpi_columns_exist():
    required = {
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "return_on_equity_pct",
        "debt_to_equity",
        "interest_coverage",
        "asset_turnover",
        "free_cash_flow_cr",
        "capex_cr",
        "earnings_per_share",
        "book_value_per_share",
        "dividend_payout_ratio_pct",
        "total_debt_cr",
        "cash_from_operations_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "composite_quality_score",
    }

    conn = get_connection()

    try:
        columns = {
            row[1]
            for row in conn.execute(
                "PRAGMA table_info(financial_ratios)"
            ).fetchall()
        }
    finally:
        conn.close()

    missing = required - columns

    assert not missing, (
        f"Missing KPI columns: {sorted(missing)}"
    )


# ============================================================
# 5. KPI COLUMNS ARE NOT COMPLETELY NULL
# ============================================================

def test_kpi_columns_not_null_only():
    columns = [
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "return_on_equity_pct",
        "debt_to_equity",
        "interest_coverage",
        "asset_turnover",
        "free_cash_flow_cr",
        "capex_cr",
        "earnings_per_share",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "composite_quality_score",
    ]

    conn = get_connection()

    try:
        total = conn.execute(
            "SELECT COUNT(*) FROM financial_ratios"
        ).fetchone()[0]

        for column in columns:
            populated = conn.execute(
                f"""
                SELECT COUNT(*)
                FROM financial_ratios
                WHERE "{column}" IS NOT NULL
                """
            ).fetchone()[0]

            assert populated > 0, (
                f"{column} is completely NULL"
            )

            assert populated <= total
    finally:
        conn.close()


# ============================================================
# 6. COMPOSITE QUALITY SCORE RANGE
# ============================================================

def test_composite_quality_score_range():
    conn = get_connection()

    try:
        invalid = conn.execute(
            """
            SELECT COUNT(*)
            FROM financial_ratios
            WHERE composite_quality_score IS NOT NULL
              AND (
                    composite_quality_score < 0
                    OR composite_quality_score > 100
                  )
            """
        ).fetchone()[0]
    finally:
        conn.close()

    assert invalid == 0


# ============================================================
# 7. DEBT TO EQUITY IS NOT NEGATIVE
# ============================================================

def test_debt_to_equity_non_negative():
    conn = get_connection()

    try:
        invalid = conn.execute(
            """
            SELECT COUNT(*)
            FROM financial_ratios
            WHERE debt_to_equity IS NOT NULL
              AND debt_to_equity < 0
            """
        ).fetchone()[0]
    finally:
        conn.close()

    assert invalid == 0


# ============================================================
# 8. PEER GROUP COUNT
# ============================================================

def test_peer_group_count():
    conn = get_connection()

    try:
        count = conn.execute(
            """
            SELECT COUNT(DISTINCT peer_group_name)
            FROM peer_groups
            """
        ).fetchone()[0]
    finally:
        conn.close()

    assert count == 11


# ============================================================
# 9. PEER PERCENTILE ROW COUNT
# ============================================================

def test_peer_percentile_row_count():
    conn = get_connection()

    try:
        count = conn.execute(
            """
            SELECT COUNT(*)
            FROM peer_percentiles
            """
        ).fetchone()[0]
    finally:
        conn.close()

    assert count == 560


# ============================================================
# 10. PEER PERCENTILES VALID RANGE
# ============================================================

def test_peer_percentiles_valid_range():
    conn = get_connection()

    try:
        invalid = conn.execute(
            """
            SELECT COUNT(*)
            FROM peer_percentiles
            WHERE percentile_rank IS NOT NULL
              AND (
                    percentile_rank < 0
                    OR percentile_rank > 1
                  )
            """
        ).fetchone()[0]
    finally:
        conn.close()

    assert invalid == 0


# ============================================================
# 11. TEN PEER METRICS
# ============================================================

def test_peer_metric_count():
    conn = get_connection()

    try:
        count = conn.execute(
            """
            SELECT COUNT(DISTINCT metric)
            FROM peer_percentiles
            """
        ).fetchone()[0]
    finally:
        conn.close()

    assert count == 10


# ============================================================
# 12. EVERY PEER GROUP HAS COMPANIES
# ============================================================

def test_every_peer_group_has_companies():
    conn = get_connection()

    try:
        rows = conn.execute(
            """
            SELECT
                peer_group_name,
                COUNT(DISTINCT company_id)
            FROM peer_groups
            GROUP BY peer_group_name
            """
        ).fetchall()
    finally:
        conn.close()

    assert len(rows) == 11

    for group_name, company_count in rows:
        assert company_count > 0, (
            f"Peer group {group_name} has no companies"
        )


# ============================================================
# 13. PEER COMPARISON WORKBOOK
# ============================================================

def test_peer_comparison_workbook():
    assert PEER_FILE.exists(), (
        "peer_comparison.xlsx does not exist"
    )

    workbook = load_workbook(
        PEER_FILE,
        read_only=True,
    )

    expected = {
        "Automobiles",
        "Consumer Finance",
        "FMCG",
        "IT Services",
        "Life Insurance",
        "Oil & Gas",
        "Pharmaceuticals",
        "Power & Utilities",
        "Private Banks",
        "Public Sector Banks",
        "Steel",
    }

    assert len(workbook.sheetnames) == 11
    assert set(workbook.sheetnames) == expected


# ============================================================
# 14. SCREENER + RADAR OUTPUTS
# ============================================================

def test_screener_and_radar_outputs():
    assert SCREENER_FILE.exists(), (
        "screener_output.xlsx does not exist"
    )

    workbook = load_workbook(
        SCREENER_FILE,
        read_only=True,
    )

    expected_sheets = {
        "quality_compounder",
        "value_pick",
        "growth_accelerator",
        "dividend_champion",
        "debt_free_blue_chip",
        "turnaround_watch",
    }

    assert set(workbook.sheetnames) == expected_sheets

    assert RADAR_DIR.exists(), (
        "Radar chart directory does not exist"
    )

    radar_files = list(
        RADAR_DIR.glob("*_radar.png")
    )

    assert len(radar_files) >= 90, (
        f"Expected at least 90 radar charts, "
        f"found {len(radar_files)}"
    )